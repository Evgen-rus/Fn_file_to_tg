"""
Формирование Excel-файла LeadRecord_FNG_YYYYMMDD_HHMMSS.xlsx из SQLite БД.

Заполняемые колонки из ТЗ:
- Телефон (мобильный) -> из БД leads.phone
- UTM_CAMPAIGN -> из БД leads.utm_campaign
- Направление -> из БД leads.direction

Инкрементальная выгрузка: скрипт запоминает максимальный внутренний row_id,
выгруженный по каждому направлению, и при следующем запуске выгружает только
новые записи.

Для направления сохраняется текущая логика подмены по доменам из utm_campaign.
"""

import os
import sys
import sqlite3
from datetime import datetime
from typing import List, Tuple, Dict
from urllib.parse import urlparse

from openpyxl import Workbook
import pytz
import requests
from dotenv import load_dotenv
import time
import logging
from email_sender import send_email_with_attachment_with_retries
from logging_setup import configure_logging


DB_FILENAME = 'baltlease_data.db'
STATE_TABLE = 'export_state'
MAX_TG_UPLOAD_BYTES = 49 * 1024 * 1024  # безопасный лимит < 50 МБ
LOGGER_NAME = 'app.export_to_excel'
TARIFF_TABLE = 'tariff_state'
# Базовый остаток по тарифу, используется только если в БД ещё нет значения
DEFAULT_TARIFF_REMAINING = 3000


HEADERS: List[str] = [
    'Телефон (мобильный)',
    'UTM_CAMPAIGN',
    'Направление',
]


# Домены, при наличии которых в UTM_CAMPAIGN направление в выгружаемом файле
# проставляется как "Мед оборудование"
DOMAINS_FOR_MED_DIRECTION: List[str] = [
    'lmed.ru',
]


# Домены, при наличии которых в UTM_CAMPAIGN направление в выгружаемом файле
# проставляется как "Сельхозтехника"
DOMAINS_FOR_AGRO_DIRECTION: List[str] = [
    'verum-agro.ru',    
]


def _utm_matches_med_domain(utm_campaign: str) -> bool:
    """
    Возвращает True, если значение utm_campaign совпадает или содержит один из доменов
    из списка DOMAINS_FOR_MED_DIRECTION. Нормализуем регистр и пытаемся корректно
    разобрать URL, чтобы сравнить как хост, так и хост+путь (для случаев вроде
    medtradegroup.ru/leasing).
    """
    if not utm_campaign:
        return False

    uc = (utm_campaign or '').strip().lower()
    domains = set(d.strip().lower() for d in DOMAINS_FOR_MED_DIRECTION)

    # Попытка распарсить как URL; если схемы нет — добавим временно
    parsed = urlparse(uc if '://' in uc else f'http://{uc}')
    host = (parsed.netloc or '').lower()
    path = (parsed.path or '').lower()
    host_plus_path = f"{host}{path}" if host else ''

    if host and host in domains:
        return True
    if host_plus_path and host_plus_path in domains:
        return True

    # Запасной вариант — простое вхождение подстроки
    return any(d in uc for d in domains)


def _utm_matches_agro_domain(utm_campaign: str) -> bool:
    """
    Возвращает True, если значение utm_campaign совпадает или содержит один из доменов
    из списка DOMAINS_FOR_AGRO_DIRECTION. Нормализуем регистр и пытаемся корректно
    разобрать URL, чтобы сравнить как хост, так и хост+путь (для случаев вроде
    verum-agro.ru/leasing).
    """
    if not utm_campaign:
        return False

    uc = (utm_campaign or '').strip().lower()
    domains = set(d.strip().lower() for d in DOMAINS_FOR_AGRO_DIRECTION)

    # Попытка распарсить как URL; если схемы нет — добавим временно
    parsed = urlparse(uc if '://' in uc else f'http://{uc}')
    host = (parsed.netloc or '').lower()
    path = (parsed.path or '').lower()
    host_plus_path = f"{host}{path}" if host else ''

    if host and host in domains:
        return True
    if host_plus_path and host_plus_path in domains:
        return True

    # Запасной вариант — простое вхождение подстроки
    return any(d in uc for d in domains)


def ensure_state_table(conn: sqlite3.Connection) -> None:
    conn.execute(
        f"""
        CREATE TABLE IF NOT EXISTS {STATE_TABLE} (
            direction TEXT PRIMARY KEY,
            last_exported_row_id INTEGER NOT NULL
        )
        """
    )
    conn.commit()


def load_last_ids(conn: sqlite3.Connection) -> Dict[str, int]:
    cur = conn.execute(f"SELECT direction, last_exported_row_id FROM {STATE_TABLE}")
    return {row[0]: int(row[1]) for row in cur.fetchall()}


def store_last_ids(conn: sqlite3.Connection, updates: Dict[str, int]) -> None:
    if not updates:
        return
    rows = [(direction, last_id) for direction, last_id in updates.items()]
    conn.executemany(
        f"""
        INSERT INTO {STATE_TABLE}(direction, last_exported_row_id)
        VALUES(?, ?)
        ON CONFLICT(direction) DO UPDATE SET last_exported_row_id=excluded.last_exported_row_id
        """,
        rows,
    )
    conn.commit()


def ensure_tariff_table(conn: sqlite3.Connection) -> None:
    """
    Создает таблицу для учета остатка по тарифу, если её нет.
    В таблице одна строка с id = 1 и полем remaining.
    """
    conn.execute(
        f"""
        CREATE TABLE IF NOT EXISTS {TARIFF_TABLE} (
            id INTEGER PRIMARY KEY CHECK(id = 1),
            remaining INTEGER NOT NULL,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    conn.commit()


def load_tariff_remaining(conn: sqlite3.Connection) -> int:
    """
    Возвращает текущий остаток по тарифу из БД.
    Если записи ещё нет — инициализирует значением DEFAULT_TARIFF_REMAINING.
    """
    ensure_tariff_table(conn)
    cur = conn.execute(f"SELECT remaining FROM {TARIFF_TABLE} WHERE id = 1")
    row = cur.fetchone()
    if row is None:
        conn.execute(f"INSERT INTO {TARIFF_TABLE}(id, remaining) VALUES (1, ?)", (DEFAULT_TARIFF_REMAINING,))
        conn.commit()
        return DEFAULT_TARIFF_REMAINING
    return int(row[0])


def save_tariff_remaining(conn: sqlite3.Connection, value: int) -> None:
    """
    Сохраняет новый остаток по тарифу (допускает отрицательное значение).
    """
    ensure_tariff_table(conn)
    safe_value = int(value)
    conn.execute(
        f"""
        INSERT INTO {TARIFF_TABLE}(id, remaining, updated_at)
        VALUES (1, ?, CURRENT_TIMESTAMP)
        ON CONFLICT(id) DO UPDATE SET remaining=excluded.remaining, updated_at=CURRENT_TIMESTAMP
        """,
        (safe_value,),
    )
    conn.commit()


def fetch_new_rows_with_state() -> Tuple[List[Tuple[str, str, str, str]], Dict[str, int], List[int]]:
    """
    Возвращает новые строки для выгрузки, словарь максимальных row_id по каждому направлению
    и список row_id всех выгруженных лидов(идентификаторов).

    Строки для Excel: (phone, utm_campaign, direction, status).
    """
    if not os.path.exists(DB_FILENAME):
        logging.getLogger(LOGGER_NAME).error(
            f"База {DB_FILENAME} не найдена. Сначала запустите экспорт в БД."
        )
        sys.exit(1)

    with sqlite3.connect(DB_FILENAME) as conn:
        ensure_state_table(conn)
        last_ids = load_last_ids(conn)

        directions = [
            row[0] for row in conn.execute("SELECT DISTINCT direction FROM leads").fetchall()
        ]

        rows: List[Tuple[str, str, str, str]] = []
        max_ids: Dict[str, int] = {}
        exported_row_ids: List[int] = []

        for direction in directions:
            last_row_id = last_ids.get(direction, 0)
            cur = conn.execute(
                """
                SELECT row_id, phone, utm_campaign, direction, status
                FROM leads
                WHERE direction = ? AND row_id > ?
                ORDER BY row_id
                """,
                (direction, last_row_id),
            )
            rows_for_direction = cur.fetchall()
            if not rows_for_direction:
                continue

            for (row_id, phone, utm, dirn, status) in rows_for_direction:
                if row_id is not None:
                    exported_row_ids.append(int(row_id))
                rows.append((phone, utm, dirn, status))

            max_row_id = max(r[0] for r in rows_for_direction if r[0] is not None)
            if max_row_id is not None:
                max_ids[direction] = max_row_id

        return rows, max_ids, exported_row_ids


def mark_rows_as_sent(conn: sqlite3.Connection, row_ids: List[int], sent_at: str) -> None:
    """
    Помечает переданные row_id как отправленные (записывает дату/время в sent_at).
    Существующие значения sent_at не перезаписываются.
    """
    if not row_ids:
        return

    placeholders = ",".join("?" for _ in row_ids)
    sql = f"""
        UPDATE leads
        SET sent_at = ?
        WHERE row_id IN ({placeholders})
          AND (sent_at IS NULL OR sent_at = '')
    """
    conn.execute(sql, (sent_at, *row_ids))
    conn.commit()


def build_workbook(rows: List[Tuple[str, str, str, str]]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Лиды'

    # Заголовки
    ws.append(HEADERS)

    # Индексы трех итоговых колонок файла
    idx_mobile = HEADERS.index('Телефон (мобильный)') + 1
    idx_utm = HEADERS.index('UTM_CAMPAIGN') + 1
    idx_direction = HEADERS.index('Направление') + 1

    for phone, utm_campaign, direction, status in rows:
        row = [''] * len(HEADERS)
        row[idx_mobile - 1] = phone or ''
        row[idx_utm - 1] = utm_campaign or ''
        # Если utm_campaign указывает на один из заданных доменов —
        # в файл выводим соответствующее направление ("Мед оборудование" или "Сельхозтехника")
        if _utm_matches_med_domain(utm_campaign or ''):
            effective_direction = 'Мед оборудование'
        elif _utm_matches_agro_domain(utm_campaign or ''):
            effective_direction = 'Сельхозтехника'
        else:
            effective_direction = direction or ''
        row[idx_direction - 1] = effective_direction
        ws.append(row)
        # Обозначим ячейку телефона как текстовую, чтобы Excel не ругался
        last_row = ws.max_row
        ws.cell(row=last_row, column=idx_mobile).number_format = '@'

    return wb


def send_text_message(token: str, chat_id: str, text: str, timeout: int = 30) -> bool:
    url = f"https://api.telegram.org/bot{token}/sendMessage"
    try:
        resp = requests.post(url, json={"chat_id": chat_id, "text": text}, timeout=timeout)
        if resp.status_code != 200:
            logging.getLogger(LOGGER_NAME).error(f"Ошибка отправки сообщения в Telegram: HTTP {resp.status_code} {resp.text}")
            return False
        data = resp.json()
        if not data.get("ok"):
            logging.getLogger(LOGGER_NAME).error(f"Telegram API вернул ошибку при отправке сообщения: {data}")
            return False
        return True
    except requests.RequestException as e:
        logging.getLogger(LOGGER_NAME).error(f"Сетевая ошибка при отправке сообщения в Telegram: {e}")
        return False


def send_document_with_retries(
    token: str,
    chat_id: str,
    file_path: str,
    caption: str | None = None,
    max_retries: int = 3,
    base_delay_sec: float = 1.0,
) -> bool:
    if not os.path.exists(file_path):
        logging.getLogger(LOGGER_NAME).error(f"Файл для отправки не найден: {file_path}")
        return False

    file_size = os.path.getsize(file_path)
    if file_size > MAX_TG_UPLOAD_BYTES:
        mb = round(file_size / (1024 * 1024), 2)
        err_text = (
            f"Не удалось отправить файл: превышен лимит Telegram (50 МБ). Размер: {mb} МБ.\n"
            f"{caption or ''}"
        ).strip()
        send_text_message(token, chat_id, err_text)
        logging.getLogger(LOGGER_NAME).error("Превышен лимит размера файла для Telegram, отправлено текстовое уведомление в чат.")
        return False

    url = f"https://api.telegram.org/bot{token}/sendDocument"
    attempt = 0
    while attempt < max_retries:
        attempt += 1
        try:
            with open(file_path, 'rb') as fh:
                files = {"document": (os.path.basename(file_path), fh)}
                data = {"chat_id": chat_id}
                if caption:
                    data["caption"] = caption
                resp = requests.post(url, data=data, files=files, timeout=60)
            if resp.status_code == 200:
                body = resp.json()
                if body.get("ok"):
                    logging.getLogger(LOGGER_NAME).info(f"Файл отправлен в Telegram (попытка {attempt}/{max_retries}).")
                    return True
                else:
                    logging.getLogger(LOGGER_NAME).error(f"Telegram API вернул ошибку: {body}")
            else:
                logging.getLogger(LOGGER_NAME).error(f"Ошибка HTTP при отправке файла: {resp.status_code} {resp.text}")
        except requests.RequestException as e:
            logging.getLogger(LOGGER_NAME).warning(f"Сетевая ошибка при отправке файла (попытка {attempt}/{max_retries}): {e}")

        if attempt < max_retries:
            delay = base_delay_sec * (2 ** (attempt - 1))
            logging.getLogger(LOGGER_NAME).info(f"Повторная попытка через {delay:.1f} сек...")
            time.sleep(delay)

    logging.getLogger(LOGGER_NAME).error("Не удалось отправить файл в Telegram после всех попыток.")
    return False


def main():
    logger = configure_logging('export_to_excel')
    load_dotenv()
    token = os.getenv('TELEGRAM_BOT_TOKEN_ASSISTANT')
    chat_id = os.getenv('TELEGRAM_CHAT_ID')

    if not token or not chat_id:
        logger.error("Не заданы переменные окружения TELEGRAM_BOT_TOKEN_ASSISTANT или TELEGRAM_CHAT_ID.")
        sys.exit(1)

    rows, max_ids, exported_row_ids = fetch_new_rows_with_state()
    if not rows:
        logger.info("Нет новых строк для выгрузки — файл не создан.")
        return

    wb = build_workbook(rows)
    ts = datetime.now(pytz.timezone('Europe/Moscow')).strftime('%Y-%m-%d_%H-%M-%S')
    filename = f"LeadRecord_FNG_{ts}.xlsx"
    wb.save(filename)
    logger.info(f"Файл сохранён: {filename}")

    # Подготовка подписи с учетом остатка по тарифу (считаем остаток до списания)
    try:
        with sqlite3.connect(DB_FILENAME) as conn:
            current_remaining = load_tariff_remaining(conn)
    except Exception:
        current_remaining = DEFAULT_TARIFF_REMAINING
    new_remaining = current_remaining - len(rows)
    caption = (
        f"Загружено новых идентификаторов: {len(rows)}\n"
        f"Остаток по тарифу: {new_remaining}"
    )

    telegram_ok = send_document_with_retries(
        token=token,
        chat_id=chat_id,
        file_path=filename,
        caption=caption,
        max_retries=3,
    )

    email_subject = f"Новые идентификаторы LeadRecord_FNG: {len(rows)}"
    email_body = (
        f"Загружено новых идентификаторов: {len(rows)}\n"
        f"Остаток по тарифу: {new_remaining}\n"
        f"Во вложении файл: {filename}"
    )
    email_ok, email_status_text = send_email_with_attachment_with_retries(
        subject=email_subject,
        body=email_body,
        attachment_path=filename,
        max_retries=5,
    )
    logger.info(
        "Итог отправки: email=%s, telegram=%s",
        "ok" if email_ok else "failed",
        "ok" if telegram_ok else "failed",
    )

    if telegram_ok:
        telegram_email_status_text = (
            email_status_text
            if email_ok
            else f"Письмо на почту не отправлено. Причина: {email_status_text}"
        )
        send_text_message(token, chat_id, telegram_email_status_text)

    # Приоритетный канал доставки — email. Если письмо отправлено успешно,
    # считаем выгрузку завершенной даже при недоступном Telegram.
    if email_ok:
        if not telegram_ok:
            logger.warning(
                "Отправка в Telegram не удалась, но письмо отправлено успешно. "
                "Состояние экспорта будет обновлено."
            )
        try:
            os.remove(filename)
            logger.info(f"Файл {filename} удалён после успешной отправки письма.")
        except Exception as e:
            logger.warning(f"Не удалось удалить файл {filename}: {e}")

        sent_at_str = datetime.now(pytz.timezone('Europe/Moscow')).strftime('%Y-%m-%d %H:%M:%S')

        with sqlite3.connect(DB_FILENAME) as conn:
            ensure_state_table(conn)
            store_last_ids(conn, max_ids)
            try:
                save_tariff_remaining(conn, new_remaining)
                logger.info(f"Остаток по тарифу обновлён: {new_remaining}")
            except Exception as e:
                logger.warning(f"Не удалось обновить остаток по тарифу: {e}")
            try:
                mark_rows_as_sent(conn, exported_row_ids, sent_at_str)
                logger.info(f"Поле sent_at обновлено для {len(exported_row_ids)} идентификаторов.")
            except Exception as e:
                logger.warning(f"Не удалось обновить sent_at для идентификаторов: {e}")
            logger.info("Состояние экспорта обновлено.")
    else:
        logger.error(
            "Письмо не отправлено. Состояние экспорта не обновлено, файл сохранён для проверки."
        )
        if not telegram_ok:
            logger.warning("Дополнительно не удалось отправить файл в Telegram.")
        sys.exit(1)


if __name__ == '__main__':
    main()


